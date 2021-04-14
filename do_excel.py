# encoding: utf-8 
"""
@author: 凹凸曼
@number: 1526
@file: do_excel.py
@time: 2019/4/20 14:51
"""
import openpyxl
# from API_6.common import http_request

class Case:
    """
    测试用例类，实例每个测试用例
    """
    def __init__(self):

        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None

class DoExcel:
    """
    操作excel类：读、写
    """
    def __init__(self,file_name,sheet_name):

        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet = self.workbook[sheet_name]

    def get_cases(self):

        max_row = self.sheet.max_row
        cases = []
        for r in range(2,max_row+1):
            case = Case()
            case.case_id = self.sheet.cell(r,1).value
            case.title = self.sheet.cell(r, 2).value
            case.url = self.sheet.cell(r, 3).value
            case.data = self.sheet.cell(r, 4).value
            case.method = self.sheet.cell(r, 5).value
            case.expected = self.sheet.cell(r, 6).value
            case.sql = self.sheet.cell(r, 9).value
            cases.append(case)
        self.workbook.close()                               #一定要关闭
        return cases

    def write_result(self, row, actual, result):

        sheet = self.workbook[self.sheet_name]
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        self.workbook.save(filename=self.file_name)         #保存
        self.workbook.close()                               #关闭

if __name__ == '__main__':

    from API_5.common import contants
    #登入
    # do_excel = DoExcel(contants.case_life, sheet_name='login')
    # cases = do_excel.get_cases()
    # http_request = http_request.HttpRequest()
    # for case in cases:
    #     # print(case.__dict__)
    #     resp = http_request.request(case.method, case.url, case.data)
    #     print(resp.status_code)
    #     print(resp.text)
        # resp_dict = resp.json()  # 返回字典
        # print(resp_dict)
        # actual = resp.text
        # if case.expected == actual:  # 判断期望结果是否与实际结果一致
        #     do_excel.write_result(case.case_id + 1, actual, 'PASS')

        # else:
        #     do_excel.write_result(case.case_id + 1, actual, 'FAIL')

    #注册
    # do_excel = DoExcel(contants.case_life, sheet_name='register')
    # cases = do_excel.get_cases()
    # http_request = http_request.HttpRequest()
    # for case in cases:
    #     resp = http_request.request(case.method, case.url, case.data)
    #     print(resp.status_code)
    #     print(resp.text)

    #充值
    # do_excel = DoExcel(contants.case_life, sheet_name='recharge')
    # cases = do_excel.get_cases()
    # http_request = http_request.HttpRequest()
    # for case in cases:
    #     resp = http_request.request(case.method, case.url, case.data)
    #     print(resp.status_code)
    #     print(case.case_id)
    #     print(resp.text)

